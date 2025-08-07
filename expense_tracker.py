import tkinter as tk
from tkinter import messagebox, ttk
import csv
import os
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

CSV_FILE = "expenses.csv"

def initialize_csv():
    """Create CSV file if it doesn't exist."""
    if not os.path.exists(CSV_FILE) or os.path.getsize(CSV_FILE) == 0:
        with open(CSV_FILE, "w", newline="") as file:
            writer = csv.writer(file)
            writer.writerow(["Date", "Type", "Amount", "Description"])

def autofill_income():
    """Autofill income entry if already stored for the month."""
    month = date_entry.get()
    if not os.path.exists(CSV_FILE):
        return
    with open(CSV_FILE, "r") as file:
        reader = csv.DictReader(file)
        for row in reader:
            if row["Date"] == month and row["Type"] == "INCOME":
                income_entry.delete(0, tk.END)
                income_entry.insert(0, row["Amount"])
                income_desc_entry.delete(0, tk.END)
                income_desc_entry.insert(0, row["Description"])
                return

def add_income():
    """Add or replace monthly income."""
    date = date_entry.get()
    amount = income_entry.get()
    desc = income_desc_entry.get()

    if not date or not amount:
        messagebox.showerror("Error", "Date and income amount are required.")
        return

    try:
        amount = float(amount)
    except ValueError:
        messagebox.showerror("Error", "Invalid income amount.")
        return

    rows = []
    # read existing and remove any existing income for this month
    with open(CSV_FILE, "r", newline="") as file:
        reader = csv.reader(file)
        header = next(reader)
        for row in reader:
            if not (row[0] == date and row[1] == "INCOME"):
                rows.append(row)

    with open(CSV_FILE, "w", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(header)
        writer.writerows(rows)
        writer.writerow([date, "INCOME", amount, desc])

    messagebox.showinfo("Success", f"Income for {date} set to ₱{amount}")
    autofill_income()
    show_summary()

def add_expense():
    """Add a new expense row."""
    date = date_entry.get()
    amount = expense_entry.get()
    desc = expense_desc_entry.get()

    if not date or not amount:
        messagebox.showerror("Error", "Date and expense amount are required.")
        return

    try:
        amount = float(amount)
    except ValueError:
        messagebox.showerror("Error", "Invalid expense amount.")
        return

    with open(CSV_FILE, "a", newline="") as file:
        writer = csv.writer(file)
        writer.writerow([date, "EXPENSE", amount, desc])

    messagebox.showinfo("Success", f"Expense of ₱{amount} added.")
    show_summary()

def show_summary():
    """Display monthly summary in the Tkinter table."""
    for item in summary_table.get_children():
        summary_table.delete(item)

    data = defaultdict(lambda: {"income": 0, "expense": 0})
    total_saved = 0.0

    # read fresh
    if not os.path.exists(CSV_FILE):
        total_saved_label.config(text=f"💰 Total Saved: ₱{0:,.2f}")
        return

    with open(CSV_FILE, "r") as file:
        reader = csv.DictReader(file)
        for row in reader:
            date = row["Date"]
            try:
                amount = float(row["Amount"]) if row["Amount"] else 0
            except ValueError:
                amount = 0
            if row["Type"] == "INCOME":
                data[date]["income"] += amount
            else:
                data[date]["expense"] += amount

    for month, values in sorted(data.items()):
        income = values["income"]
        expense = values["expense"]
        remaining = income - expense
        total_saved += remaining
        summary_table.insert("", "end", values=(
            month, f"₱{income:,.2f}", f"₱{expense:,.2f}", f"₱{remaining:,.2f}"
        ))

    total_saved_label.config(text=f"💰 Total Saved: ₱{total_saved:,.2f}")

def export_to_excel():
    """Export monthly summary + detailed expenses to Excel with separate sheets."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Styles
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    align = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True)

    # Load data
    month_data = defaultdict(list)
    summary_data = defaultdict(lambda: {"income": 0, "expense": 0})

    with open(CSV_FILE, "r") as file:
        reader = csv.DictReader(file)
        for row in reader:
            month = row["Date"]
            amount = float(row["Amount"])
            month_data[month].append(row)
            if row["Type"] == "INCOME":
                summary_data[month]["income"] += amount
            elif row["Type"] == "EXPENSE":
                summary_data[month]["expense"] += amount

    # Create a sheet for each month
    for month, rows in sorted(month_data.items()):
        ws = wb.create_sheet(title=month)

        # --- Summary Section ---
        headers = ["Month", "Income", "Expenses", "Remaining"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = bold_font
            cell.alignment = align
            cell.border = border
            cell.fill = header_fill

        income = summary_data[month]["income"]
        expense = summary_data[month]["expense"]
        remaining = income - expense

        ws.append([month, income, expense, remaining])
        ws.cell(row=2, column=2).fill = yellow_fill
        ws.cell(row=2, column=3).fill = red_fill
        ws.cell(row=2, column=4).fill = green_fill
        for col in range(1, 5):
            cell = ws.cell(row=2, column=col)
            cell.border = border
            cell.alignment = align

        # --- Expense List Section ---
        start_row = 4
        exp_headers = ["Date", "Type", "Amount", "Description"]
        for col_num, header in enumerate(exp_headers, 1):
            cell = ws.cell(row=start_row, column=col_num)
            cell.value = header
            cell.font = bold_font
            cell.alignment = align
            cell.border = border
            cell.fill = header_fill

        for r_idx, expense in enumerate(rows, start=start_row + 1):
            ws.cell(row=r_idx, column=1, value=expense["Date"])
            ws.cell(row=r_idx, column=2, value=expense["Type"])
            ws.cell(row=r_idx, column=3, value=float(expense["Amount"]))
            ws.cell(row=r_idx, column=4, value=expense["Description"])
            for col in range(1, 5):
                cell = ws.cell(row=r_idx, column=col)
                cell.border = border
                cell.alignment = align

    wb.save("Monthly_Salary_with_Expenses.xlsx")
    messagebox.showinfo("Exported", "Detailed monthly salary file saved as 'Monthly_Salary_with_Expenses.xlsx'")



# --- GUI Setup ---
initialize_csv()

root = tk.Tk()
root.title("📊 Monthly Expense Tracker")
root.geometry("600x650")
root.configure(bg="#f0f0f0")

# Date
tk.Label(root, text="Month (YYYY-MM):", bg="#f0f0f0").pack()
date_entry = tk.Entry(root)
date_entry.insert(0, datetime.now().strftime("%Y-%m"))
date_entry.pack()
date_entry.bind("<FocusOut>", lambda e: autofill_income())

# Income section
tk.Label(root, text="\n💵 Income", font=("Arial", 12, "bold"), bg="#f0f0f0").pack()
tk.Label(root, text="Amount:", bg="#f0f0f0").pack()
income_entry = tk.Entry(root)
income_entry.pack()
tk.Label(root, text="Description:", bg="#f0f0f0").pack()
income_desc_entry = tk.Entry(root)
income_desc_entry.pack()
tk.Button(root, text="Add Income", command=add_income, bg="lightgreen", font=("Arial", 10, "bold")).pack(pady=5)

# Expense section
tk.Label(root, text="\n🧾 Expense", font=("Arial", 12, "bold"), bg="#f0f0f0").pack()
tk.Label(root, text="Amount:", bg="#f0f0f0").pack()
expense_entry = tk.Entry(root)
expense_entry.pack()
tk.Label(root, text="Description:", bg="#f0f0f0").pack()
expense_desc_entry = tk.Entry(root)
expense_desc_entry.pack()
tk.Button(root, text="Add Expense", command=add_expense, bg="lightcoral", font=("Arial", 10, "bold")).pack(pady=5)

# Table View
tk.Label(root, text="\n📅 Monthly Summary", font=("Arial", 12, "bold"), bg="#f0f0f0").pack()

style = ttk.Style()
style.configure("Treeview.Heading", font=("Arial", 10, "bold"), background="#dcdcdc")
style.configure("Treeview", rowheight=15)

columns = ("Month", "Income", "Expenses", "Remaining")
summary_table = ttk.Treeview(root, columns=columns, show="headings", height=6)
for col in columns:
    summary_table.heading(col, text=col)
    summary_table.column(col, anchor="center", width=130)
summary_table.pack(pady=10)

total_saved_label = tk.Label(root, text="💰 Total Saved: ₱0.00", font=("Arial", 11, "bold"), bg="#f0f0f0")
total_saved_label.pack()

tk.Button(root, text="📤 Export to Excel", command=export_to_excel, bg="#4da6ff", fg="white", font=("Arial", 10, "bold")).pack(pady=10)

show_summary()
root.mainloop()
